#!/usr/bin/env python3
"""
Brand Management System for World-Class Carousel
=================================================
Axiom A1: A brand is a consistent perceptual promise.
A viewer must identify the brand from any single slide without reading the name.

Architecture:
  brands/{slug}/brand.json          -- Brand identity + design language
  brands/{slug}/logo.png            -- AI-generated brand logo
  brands/{slug}/content_log.xlsx    -- Topics covered, dates, metrics
  brands/{slug}/design_language.json -- Full 7-axis spec for renderer

Usage:
  from brand_manager import BrandManager
  bm = BrandManager()
  bm.list_brands()
  brand = bm.load_brand("my-brand")
  bm.check_topic_exists(brand, "GPT-5 vs Claude comparison")
  bm.log_content(brand, "GPT-5 vs Claude", "tool_showdown", 8)
"""

import json, os, re
from pathlib import Path
from datetime import datetime

BRANDS_DIR = Path("/home/node/.claude/skills/world-class-carousel/brands")
BRANDS_DIR.mkdir(parents=True, exist_ok=True)

# Available installed fonts (verified working with LaTeX)
INSTALLED_FONTS = [
    "newpxtext",    # Palatino serif
    "newtxtext",    # Times serif
    "charter",      # Charter serif
    "bookman",      # Bookman serif
    "helvet",       # Helvetica sans
    "roboto",       # Roboto sans
    "cabin",        # Cabin humanist sans
    "opensans",     # Open Sans
    "avant",        # Avant Garde geometric sans
    "inconsolata",  # Inconsolata mono
]

# =====================================================
# DERIVATION TABLE: Interview Answers -> 7-Axis Values
# =====================================================
# Maps 4 interview dimensions to 7 design language axes

TONE_MAP = {
    "authoritative": {
        "font": "newpxtext",       # Serif = authority
        "density": "balanced",
        "depth": "flat",
        "header_style": "italic",
        "divider_style": "ornament",
    },
    "friendly": {
        "font": "cabin",           # Humanist sans = warmth
        "density": "generous",
        "depth": "subtle",
        "header_style": "bold",
        "divider_style": "dots",
    },
    "edgy": {
        "font": "avant",           # Heavy geometric = confrontation
        "density": "medium",
        "depth": "flat",
        "header_style": "bold",
        "divider_style": "none",
    },
    "premium": {
        "font": "opensans",        # Light clean = Apple
        "density": "airy",
        "depth": "glass",
        "header_style": "plain",
        "divider_style": "none",
    },
    "playful": {
        "font": "roboto",          # Round modern = fun
        "density": "generous",
        "depth": "subtle",
        "header_style": "bold",
        "divider_style": "dots",
    },
}

NICHE_MAP = {
    "ai_tech": {
        "accent": "7C3AED",        # Purple
        "accent_light": "A78BFA",
        "accent_secondary": "2DD4BF",
        "bg_pattern": "circuit",
        "label_bg": "7C3AED",
    },
    "business": {
        "accent": "166534",        # Money green
        "accent_light": "4ADE80",
        "accent_secondary": "CA8A04",
        "bg_pattern": "none",
        "label_bg": "166534",
    },
    "dev_tools": {
        "accent": "0D9488",        # Teal
        "accent_light": "5EEAD4",
        "accent_secondary": "3B82F6",
        "bg_pattern": "grid",
        "label_bg": "0D9488",
    },
    "news": {
        "accent": "DC2626",        # Red urgency
        "accent_light": "F87171",
        "accent_secondary": "F59E0B",
        "bg_pattern": "scanlines",
        "label_bg": "DC2626",
    },
    "education": {
        "accent": "2563EB",        # Trust blue
        "accent_light": "93C5FD",
        "accent_secondary": "16A34A",
        "bg_pattern": "dots",
        "label_bg": "2563EB",
    },
    "creative": {
        "accent": "DB2777",        # Pink/magenta
        "accent_light": "F472B6",
        "accent_secondary": "F59E0B",
        "bg_pattern": "none",
        "label_bg": "DB2777",
    },
}

ENERGY_MAP = {
    "dark_dramatic": {
        "bg_fallback": "0D1117",
        "dark": "E6EDF3",      # Light text on dark
        "mid": "8B949E",
        "light": "30363D",
        "card_bg": "161B22",
        "corner_radius": "3pt",
        "layout": "asymmetric",
        "rhythm": "organic",
    },
    "light_airy": {
        "bg_fallback": "FAFAFA",
        "dark": "18181B",      # Dark text on light
        "mid": "71717A",
        "light": "E4E4E7",
        "card_bg": "FFFFFF",
        "corner_radius": "8pt",
        "layout": "center",
        "rhythm": "grid",
    },
    "warm_inviting": {
        "bg_fallback": "FDF6E3",
        "dark": "3C2415",
        "mid": "78716C",
        "light": "E7E0D5",
        "card_bg": "FEF9EF",
        "corner_radius": "5pt",
        "layout": "left",
        "rhythm": "organic",
    },
    "bold_contrast": {
        "bg_fallback": "0A0A14",
        "dark": "F0F0F5",
        "mid": "9CA3AF",
        "light": "1E1E2E",
        "card_bg": "141420",
        "corner_radius": "0pt",
        "layout": "asymmetric",
        "rhythm": "alternating",
    },
}


def slugify(name):
    """Convert brand name to filesystem-safe slug."""
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')


def derive_design_language(tone, niche, energy, brand_name):
    """
    Derive a complete 7-axis design language from interview answers.
    Maps: tone -> typography+density+depth, niche -> color+decoration, energy -> layout+rhythm+bg.
    """
    t = TONE_MAP.get(tone, TONE_MAP["friendly"])
    n = NICHE_MAP.get(niche, NICHE_MAP["ai_tech"])
    e = ENERGY_MAP.get(energy, ENERGY_MAP["dark_dramatic"])

    return {
        "theme": {
            "bg_fallback": e["bg_fallback"],
            "accent": n["accent"],
            "accent_light": n["accent_light"],
            "dark": e["dark"],
            "mid": e["mid"],
            "light": e["light"],
            "label_bg": n["label_bg"],
            "label_text": "FFFFFF",
            "card_bg": e["card_bg"],
        },
        "brand": {
            "font_serif": t["font"],
            "corner_radius": e["corner_radius"],
            "header_style": t["header_style"],
            "divider_style": t["divider_style"],
            "nav_style": "circle",
            "name": brand_name,
        },
        "bg_pattern": n["bg_pattern"],
        "accent_secondary": n.get("accent_secondary", ""),
        "bullet_style": "default",
        "title_treatment": "highlight",
        "depth": t["depth"],
        "density": t["density"],
        "layout": e["layout"],
        "rhythm": e["rhythm"],
    }


def create_brand(name, handle, tagline, niche, tone, energy, audience="general"):
    """
    Create a new brand with derived design language.
    Returns the brand dict and saves to disk.
    """
    slug = slugify(name)
    brand_dir = BRANDS_DIR / slug
    brand_dir.mkdir(parents=True, exist_ok=True)

    dl = derive_design_language(tone, niche, energy, name)

    brand = {
        "slug": slug,
        "name": name,
        "handle": handle,
        "tagline": tagline,
        "niche": niche,
        "tone": tone,
        "energy": energy,
        "audience": audience,
        "design_language": dl,
        "logo_path": f"brands/{slug}/logo.png",
        "content_log_path": f"brands/{slug}/content_log.xlsx",
        "created_date": datetime.now().strftime("%Y-%m-%d"),
        "categories_enabled": _default_categories(niche),
    }

    # Save brand.json
    (brand_dir / "brand.json").write_text(json.dumps(brand, indent=2))

    # Save design_language.json (renderer-compatible)
    (brand_dir / "design_language.json").write_text(json.dumps(dl, indent=2))

    return brand


def _default_categories(niche):
    """Map niche to relevant content categories."""
    base = ["breaking_news", "hot_take", "future_scenario"]
    niche_cats = {
        "ai_tech":    ["paper_decoder", "tool_showdown", "tool_tutorial", "prompt_playbook", "industry_map"],
        "business":   ["founders_money", "industry_map", "build_this"],
        "dev_tools":  ["tool_showdown", "tool_tutorial", "build_this", "prompt_playbook"],
        "news":       ["breaking_news", "industry_map", "founders_money"],
        "education":  ["paper_decoder", "tool_tutorial", "prompt_playbook"],
        "creative":   ["tool_tutorial", "build_this", "prompt_playbook"],
    }
    return list(set(base + niche_cats.get(niche, [])))


def list_brands():
    """List all known brands."""
    brands = []
    for d in sorted(BRANDS_DIR.iterdir()):
        brand_file = d / "brand.json"
        if brand_file.exists():
            brand = json.loads(brand_file.read_text())
            brands.append(brand)
    return brands


def load_brand(slug):
    """Load a brand by slug."""
    brand_file = BRANDS_DIR / slug / "brand.json"
    if not brand_file.exists():
        raise FileNotFoundError(f"Brand '{slug}' not found at {brand_file}")
    return json.loads(brand_file.read_text())


def get_content_log_path(brand):
    """Get absolute path to content log xlsx."""
    return BRANDS_DIR / brand["slug"] / "content_log.xlsx"


def check_topic_exists(brand, topic_query):
    """
    Check if a topic has already been covered.
    Returns list of similar past topics (fuzzy match on keywords).
    """
    log_path = get_content_log_path(brand)
    if not log_path.exists():
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(log_path))
        ws = wb.active
        past_topics = []
        query_words = set(topic_query.lower().split())

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:  # Column B = Topic
                topic = str(row[1])
                topic_words = set(topic.lower().split())
                overlap = len(query_words & topic_words)
                if overlap >= 2:  # At least 2 shared words
                    past_topics.append({
                        "date": str(row[0]) if row[0] else "",
                        "topic": topic,
                        "category": str(row[2]) if row[2] else "",
                        "overlap": overlap,
                    })

        return sorted(past_topics, key=lambda x: -x["overlap"])
    except Exception:
        return []


def log_content(brand, topic, category, slide_count, notes=""):
    """Log a piece of content to the brand's Excel tracker."""
    log_path = get_content_log_path(brand)

    try:
        import openpyxl
        if log_path.exists():
            wb = openpyxl.load_workbook(str(log_path))
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Content Log"
            ws.append(["Date", "Topic", "Category", "Slides", "Status", "Notes"])
            # Style header
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)

        ws.append([
            datetime.now().strftime("%Y-%m-%d"),
            topic,
            category,
            slide_count,
            "published",
            notes,
        ])

        wb.save(str(log_path))
        return True
    except ImportError:
        # Fallback: use CSV if openpyxl not available
        csv_path = log_path.with_suffix(".csv")
        import csv
        write_header = not csv_path.exists()
        with open(csv_path, "a", newline="") as f:
            w = csv.writer(f)
            if write_header:
                w.writerow(["Date", "Topic", "Category", "Slides", "Status", "Notes"])
            w.writerow([
                datetime.now().strftime("%Y-%m-%d"),
                topic, category, slide_count, "published", notes
            ])
        return True


def get_brand_summary(brand):
    """Get a human-readable summary of a brand."""
    dl = brand["design_language"]
    return {
        "name": brand["name"],
        "handle": brand["handle"],
        "niche": brand["niche"],
        "tone": brand["tone"],
        "accent_color": f"#{dl['theme']['accent']}",
        "font": dl["brand"]["font_serif"],
        "corner_radius": dl["brand"]["corner_radius"],
        "bg_style": dl["theme"]["bg_fallback"],
        "depth": dl["depth"],
        "categories": brand["categories_enabled"],
        "created": brand["created_date"],
    }


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Brand Manager CLI")
        print("  list              -- List all brands")
        print("  create            -- Interactive brand creation")
        print("  info <slug>       -- Show brand info")
        print("  topics <slug>     -- Show covered topics")
        sys.exit(0)

    cmd = sys.argv[1]

    if cmd == "list":
        brands = list_brands()
        if not brands:
            print("No brands found. Create one first.")
        for b in brands:
            s = get_brand_summary(b)
            print(f"  {s['name']} ({s['handle']}) -- {s['niche']}, {s['tone']}, accent {s['accent_color']}")

    elif cmd == "info":
        slug = sys.argv[2] if len(sys.argv) > 2 else ""
        brand = load_brand(slug)
        s = get_brand_summary(brand)
        for k, v in s.items():
            print(f"  {k}: {v}")

    elif cmd == "derive-test":
        # Quick derivation test
        for tone in ["authoritative", "friendly", "edgy", "premium"]:
            for niche in ["ai_tech", "business", "dev_tools"]:
                dl = derive_design_language(tone, niche, "dark_dramatic", "Test")
                print(f"  {tone}/{niche}: accent=#{dl['theme']['accent']}, font={dl['brand']['font_serif']}, r={dl['brand']['corner_radius']}, depth={dl['depth']}")
